#*-* encoding=UTF-8 *-*

__author__ = 'daniu'

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles.colors import RED, YELLOW, BLUE, BLACK
import math
import os

class examaitonClass():
    retcode = 1 # 0-成功；1-失败
    retmsg = ''
    dataList = []
    dataListW = []
    dataListL = []

    def __init__(self, name, fileName, num, wl):
        self.name = name
        self.fileName = fileName
        self.num = num
        self.wl = wl
        self.path = os.path.split(self.fileName)[0]
        self.dataList = []
        self.dataListW = []
        self.dataListL = []



    def main(self):

        try:
            try:
                wb = load_workbook(self.fileName)

                sheet_names = wb.get_sheet_names()
                ws = wb.get_sheet_by_name(sheet_names[0])
            except Exception as e:
                raise Exception('请选择正确的成绩单文件：' + self.fileName)

            # 读取数据
            if self.readGrade(ws):
                # 根据总分排倒叙
                self.dataList = sorted(self.dataList, key=lambda x:(x.__getitem__('zf')), reverse=True)

                # 考场分布文件
                if self.examFile():

                    # 班级分布文件
                    if self.bjFile():

                        # 考场考号文件
                        if self.kckhFile():
                            self.retcode = 0

        except Exception as e:
            self.retmsg = e
        finally:
            try:
                wb.close()
            except Exception as e:
                print(e)

        return self.retcode, self.retmsg


    # 文件1 考场分布
    def examFile(self):
        try:
            # 文件信息
            examName = os.path.join(self.path, self.name + '-考场分布.xlsx')
            print(examName)
            examRow2Title  = ['考号','姓名','班级','考场','座位号']
            lExamNums = 0
            wExamNums = 0
            examNums = 0

            dataListExam = self.dataList
            dataListExamLen = len(dataListExam)


            # 格式
            font1 = Font(name=u'宋体', size=18)
            font2 = Font(name=u'宋体', size=16)
            align = Alignment(horizontal='center', vertical='center')
            thin = Side(border_style='thin', color=BLACK)
            border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # 文理分开计算
            if self.wl == '是':
                self.dataListL = list(filter(lambda x:x.__getitem__('wl') == '理', dataListExam))
                self.dataListL = sorted(self.dataListL, key=lambda x:(x.__getitem__('zf')), reverse=True)
                lExamNums = math.ceil(len(self.dataListL)/int(self.num))
                if len(self.dataListL) == 0:
                    lExamNums = 0

                self.dataListW = list(filter(lambda x: x.__getitem__('wl') == '文', dataListExam))
                self.dataListW = sorted(self.dataListW, key=lambda x: (x.__getitem__('zf')), reverse=True)
                wExamNums = math.ceil(len(self.dataListW)/int(self.num))
                if len(self.dataListW) == 0:
                    wExamNums = 0
                len(self.dataListW)
            else:
                examNums = math.ceil(len(dataListExam)/int(self.num))
            try:
                try:
                    # 创建工作簿
                    wb_exam = Workbook()

                    # 创建工作表
                    if self.wl == '是':
                        dataListLLen = len(self.dataListL)
                        dataListWLen = len(self.dataListW)

                        # 开始写理科
                        sheetList = []
                        for i in range(0, lExamNums):
                            sheetList.append(str(i + 1) + '考场')
                            wb_exam.create_sheet(index=i, title=str(i + 1) + '考场')

                        # 开始循环考场
                        kcnum = 1  # 初始化考场
                        for sheetName in sheetList:
                            
                            ws = wb_exam.get_sheet_by_name(sheetName)

                            # 考场编号
                            ws.merge_cells('A1:E1')
                            ws.cell(row=1, column=1).value = '第' + sheetName

                            # 标题
                            ws.append(examRow2Title)

                            # 写入考场数据
                            # 计算本考场人员在datalist中下标
                            pindex_min = kcnum * int(self.num) - int(self.num)
                            pindex_max = kcnum * int(self.num) - 1
                            kcnum += 1

                            if pindex_max >= dataListLLen:
                                pindex_max = dataListLLen - 1
                            stuList = []
                            jsq = 0
                            for j in range(pindex_min, pindex_max + 1):
                                jsq += 1
                                stu = self.dataListL[j]
                                stuList.append([stu['kh'], stu['xm'], stu['bj'], sheetName, str(jsq)])


                            # 写入
                            for k in range(0, len(stuList)):
                                ws.append(stuList[k])

                            self.setExamCellStype(ws, dataListLLen, 'exam')

                        # 开始写文科
                        # if dataListWLen != 0:
                        sheetList = []
                        for i in range(lExamNums, lExamNums+wExamNums):
                            sheetList.append(str(i+1) + '考场')
                            wb_exam.create_sheet(index=i, title=str(i+1) + '考场')

                        # 开始循环考场
                        kcnum = 1  # 初始化考场
                        for sheetName in sheetList:
                            ws = wb_exam.get_sheet_by_name(sheetName)

                            # 考场编号
                            ws.merge_cells('A1:E1')
                            ws.cell(row=1, column=1).value = '第' + sheetName

                            # 标题
                            ws.append(examRow2Title)

                            # 写入考场数据
                            # 计算本考场人员在datalist中下标
                            pindex_min = kcnum * int(self.num) - int(self.num)
                            pindex_max = kcnum * int(self.num) - 1
                            kcnum += 1

                            if pindex_max >= dataListWLen:
                                pindex_max = dataListWLen - 1

                            stuList = []
                            jsq = 0
                            for j in range(pindex_min, pindex_max + 1):
                                jsq += 1
                                stu = self.dataListW[j]
                                stuList.append([stu['kh'], stu['xm'], stu['bj'], sheetName, str(jsq)])

                            # 写入
                            for k in range(0, len(stuList)):
                                ws.append(stuList[k])

                            self.setExamCellStype(ws, dataListWLen, 'exam')
                    else:
                        sheetList = []
                        for i in range(0, examNums):
                            sheetList.append(str(i+1)+'考场')
                            wb_exam.create_sheet(index=i, title=str(i+1)+'考场')

                        # 开始循环考场
                        kcnum =1 # 初始化考场
                        for sheetName in sheetList:
                            ws = wb_exam.get_sheet_by_name(sheetName)

                            # 考场编号
                            ws.merge_cells('A1:E1')
                            ws.cell(row=1, column=1).value = '第' + sheetName

                            # 标题
                            ws.append(examRow2Title)

                            # 写入考场数据
                            # 计算本考场人员在datalist中下标
                            pindex_min = kcnum*int(self.num) - int(self.num)
                            pindex_max = kcnum*int(self.num) - 1
                            kcnum += 1

                            if pindex_max > dataListExamLen:
                                pindex_max = dataListExamLen - 1

                            stuList = []
                            jsq = 0
                            for j in range(pindex_min, pindex_max+1):
                                jsq += 1
                                stu = dataListExam[j]
                                stuList.append([stu['kh'], stu['xm'], stu['bj'], sheetName, str(jsq)])

                            # 写入
                            for k in range(0, len(stuList)):
                                ws.append(stuList[k])

                            self.setExamCellStype(ws, dataListExamLen, 'exam')
                except Exception as e:
                    raise Exception('创建考场分布文件失败：' + e)
                try:
                    wb_exam.save(examName)
                except Exception as e:
                    raise Exception('请先关闭文件：' + examName)
            except Exception as e:
                self.retmsg = e
                return False
        except Exception as e:
            print(e)
            self.retmsg = e
            return False
        finally:
            try:
                wb_exam.close()
            except Exception as e:
                print(e)
        return True

    # 文件2 班级分布
    def bjFile(self):
        try:
            # 以考场分布数据为基础操作
            self.dataList = []
            examName = os.path.join(self.path, self.name + '-考场分布.xlsx')
            bjName = os.path.join(self.path, self.name + '-班级考场安排.xlsx')
            bjRow2Title = ['考号', '姓名', '班级', '考场', '座位号']

            wb_exam_t = load_workbook(examName)
            sheet_names = wb_exam_t.get_sheet_names()

            for sheet_name in sheet_names:
                ws = wb_exam_t.get_sheet_by_name(sheet_name)

                # 获取最大行数、列数
                max_row = ws.max_row

                # 获取ws数据
                for i in range(3, max_row + 1):
                    sub_data = {}
                    sub_data['kh'] = ws.cell(i, 1).value
                    sub_data['xm'] = ws.cell(i, 2).value
                    sub_data['bj'] = ws.cell(i, 3).value
                    sub_data['kc'] = ws.cell(i, 4).value
                    sub_data['zwh'] = ws.cell(i, 5).value
                    self.dataList.append(sub_data)

             # 开始分析数据
            try:
                bjList = []
                for data in self.dataList:
                    bj = data['bj']
                    if str(bj) not in bjList:
                        bjList.append(str(bj))
                bjList = sorted(bjList)

                # 创建班级 sheet
                wb_bj = Workbook()
                for j in range(0, len(bjList)):
                    wb_bj.create_sheet(index=j, title=str(bjList[j]) + '班')
                # 为每个sheet(班级)增加标题

                for k in range(0, len(bjList)):
                    sheetName = str(bjList[k])+'班'
                    ws = wb_bj.get_sheet_by_name(sheetName)

                    # 班级编号
                    ws.merge_cells('A1:E1')
                    ws.cell(row=1, column=1).value = sheetName

                    # 标题
                    ws.append(bjRow2Title)

                # 遍历所有数据
                listNameList = []
                for listV in range(0, len(bjList)):
                    listName = 'list' + str(listV)

                    listName = []
                    listNameList.append(listName)

                for dataItem in self.dataList:
                    listD = []
                    listD.append(dataItem['kh'])
                    listD.append(dataItem['xm'])
                    listD.append(dataItem['bj'])
                    listD.append(dataItem['kc'])
                    listD.append(dataItem['zwh'])
                    for listT in range(0, len(bjList)):
                        if dataItem['bj'] == bjList[listT]:
                            listNameList[listT].append(listD)

                # 按考场、座位号，为班级排升序
                for listName in listNameList:
                    sorted(listName, key=lambda l:(l[3],l[4]))


                # 写入shell
                for x in range(0, len(bjList)):
                    sheetName = str(bjList[x])+'班'
                    ws = wb_bj.get_sheet_by_name(sheetName)

                    for y in range(0, len(listNameList[x])):
                        ws.append(listNameList[x][y])

                    # 格式
                    self.setExamCellStype(ws, len(listNameList[x]), 'bj')
            except Exception as e:
                raise Exception('创建班级考场安排失败：' + bjName)
            try:
                wb_bj.save(bjName)
            except Exception as e:
                raise Exception('请先关闭文件：' + bjName)
        except Exception as e:
            print(e)
            self.retmsg = e
            return False
        finally:
            try:
                wb_exam_t.close()
                wb_bj.close()
            except Exception as e:
                print(e)

        return True

    # 文件3 考场考号
    def kckhFile(self):
        try:
            # 以考场分布数据为基础操作
            self.dataList = []
            examName = os.path.join(self.path, self.name + '-考场分布.xlsx')
            kckhName = os.path.join(self.path, self.name + '-考场考号.xlsx')

            wb_exam_t = load_workbook(examName)
            sheet_names = wb_exam_t.get_sheet_names()

            for sheet_name in sheet_names:
                ws = wb_exam_t.get_sheet_by_name(sheet_name)

                # 获取最大行数、列数
                max_row = ws.max_row

                # 获取ws数据
                for i in range(3, max_row + 1):
                    sub_data = {}
                    sub_data['kh'] = ws.cell(i, 1).value
                    sub_data['xm'] = ws.cell(i, 2).value
                    sub_data['kc'] = ws.cell(i, 4).value
                    sub_data['zwh'] = ws.cell(i, 5).value
                    self.dataList.append(sub_data)

            # 开始分析数据
            try:
                dataListLen = len(self.dataList)
                kckhList = sheet_names
                kckhList = sorted(set(kckhList))

                # 创建 sheet
                wb_kckh = Workbook()
                for j in range(0, len(kckhList)):
                    wb_kckh.create_sheet(index=j, title=str(kckhList[j]))

                # 为每个sheet增加标题
                for k in range(0, len(kckhList)):
                    ws = wb_kckh.get_sheet_by_name(kckhList[k])

                    ws.merge_cells('A1:C1')
                    ws.merge_cells('E1:G1')
                    ws.cell(row=1, column=1).value = '第' + str(kckhList[k])
                    ws.cell(row=1, column=5).value = '第' + str(kckhList[k])

                    # 标题
                    ws.merge_cells('A2:G2')

                # # 遍历所有数据
                listNameList = []
                for listV in range(0, len(kckhList)):
                    listName = 'list' + str(listV)
                    listName = []
                    listNameList.append(listName)

                for dataItem in self.dataList:
                    listD = []
                    listD.append(dataItem['kh'])
                    listD.append(dataItem['xm'])
                    listD.append(dataItem['zwh'])
                    for listT in range(0, len(kckhList)):
                        if dataItem['kc'] == kckhList[listT]:
                            listNameList[listT].append(listD)

                # 座位号排升序
                for listName in listNameList:
                    sorted(listName, key=lambda l: l[2])

                # 处理数据
                for l in range(0, len(listNameList)):
                    listL = listNameList[l]
                    minNum = math.floor(int((self.num))/2)
                    if len(listL) >= minNum:
                        for ll in range(minNum, len(listL)):
                            index = ll - minNum
                            listL[index].append('')
                            listL[index].append(listL[ll][0])
                            listL[index].append(listL[ll][1])
                            listL[index].append(listL[ll][2])

                    maxnum = 1
                    if len(listL) >= minNum:
                        maxnum = minNum
                    else:
                        maxnum = len(listL)
                    listNameList[l] = listL[0:maxnum]

                # 写入shell
                for x in range(0, len(kckhList)):
                    sheetName = str(kckhList[x])
                    ws = wb_kckh.get_sheet_by_name(sheetName)

                    for y in range(0, len(listNameList[x])):
                        ws.append(listNameList[x][y])

                    # 格式
                    self.setExamCellStype2(ws, len(listNameList[x]))

            except Exception as e:
                raise Exception('创建考场考号失败：' + kckhName)
            try:
                wb_kckh.save(kckhName)
            except Exception as e:
                raise Exception('请先关闭文件：' + kckhName)
        except Exception as e:
            self.retmsg = e
            return False
        finally:
            try:
                wb_exam_t.close()
                wb_kckh.close()
            except Exception as e:
                print(e)

        return True

    # 读取数据
    def readGrade(self, ws):
        try:
            # 获取最大行数、列数
            max_row = ws.max_row
            max_col = ws.max_column

            if max_row < 1:
                raise Exception('成绩单不可为空：' + self.fileName)

            # 校验标题一致性
            firstRow = ['考号', '姓名', '班级', '总分', '文理科']
            firstRowData = self.get_row_value(ws, 1)
            if firstRow != firstRowData or max_col > 5:
                raise Exception('成绩单列名必须为[考号、姓名、班级、总分、文理科]，不可更改：' + self.fileName)

            # 校验文理科列合规性
            wlCol = []
            wlColData = self.get_col_value(ws, 5)
            wnum= len(list(filter(lambda x: x == '文', wlColData)))
            lnum = len(list(filter(lambda x: x == '理', wlColData)))
            if wnum + lnum + 1 != max_row:
                raise Exception('文理科列内容必须为[文或理]：' + self.fileName)

            try:
                # 获取全部数据
                for i in range(2, max_row+1):
                    sub_data={}
                    sub_data['kh'] = str(ws.cell(i, 1).value)
                    sub_data['xm'] = ws.cell(i, 2).value
                    sub_data['bj'] = str(ws.cell(i, 3).value)
                    sub_data['zf'] = ws.cell(i, 4).value
                    sub_data['wl'] = ws.cell(i, 5).value
                    self.dataList.append(sub_data)
            except Exception as e:
                raise Exception('请确认成绩单内容是否合规：' + self.fileName)
            if len(self.dataList) < 1:
                raise Exception('成绩单不可为空：' + self.fileName)
        except Exception as e:
            self.retmsg = e
            return False

        return True


    # 获取某行所有值
    def get_row_value(self, ws, row):
        columns = ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 获取某列所有值
    def get_col_value(self, ws, col):
        rows = ws.max_row
        col_data = []
        for i in range(1, rows + 1):
            cell_value = ws.cell(row=i, column=col).value
            col_data.append(cell_value)
        return col_data

    # 设置考场分布、班级单元格格式
    def setExamCellStype(self, ws, dataLen, type):
        # 格式
        fontExam = Font(name=u'宋体', size=18)
        fontExam2 = Font(name=u'宋体', size=16)
        fontExam3 = Font(name=u'宋体', size=16)

        fontBj = Font(name=u'宋体', size=18)
        fontBj2 = Font(name=u'宋体', size=16)
        fontBj3 = Font(name=u'宋体', size=16)

        fontKckh = Font(name=u'宋体', size=18)
        fontKckh2 = Font(name=u'宋体', size=16)
        fontKckh3 = Font(name=u'宋体', size=16)
        align = Alignment(horizontal='center', vertical='center')
        thin = Side(border_style='thin', color=BLACK)
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.rows:
            for cell in row:
                cell.alignment = align
                cell.border = border
                if cell.row == 1:
                    if type == 'exam':
                        cell.font = fontExam
                    elif type == 'bj':
                        cell.font = fontBj
                    else:
                        cell.font = fontKckh
                elif cell.row == 2:
                    if type == 'exam':
                        cell.font = fontExam2
                    elif type == 'bj':
                        cell.font = fontBj2
                    else:
                        cell.font = fontKckh2
                else:
                    if type == 'exam':
                        cell.font = fontExam3
                    elif type == 'bj':
                        cell.font = fontBj3
                    else:
                        cell.font = fontKckh3

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13

        for rowX in range(1, dataLen + 1):
            ws.row_dimensions[rowX].height = 21

    # 设置考场考号单元格格式
    def setExamCellStype2(self, ws, dataLen):
        # 格式
        fontKckh = Font(name=u'宋体', size=16)
        fontKckh2 = Font(name=u'宋体', size=20)
        align = Alignment(horizontal='center', vertical='center')
        thin = Side(border_style='thin', color=BLACK)
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for col in ws.columns:
            for cell in col:
                if cell.column != 4:
                    cell.alignment = align
                    cell.border = border
                if cell.row == 1:
                    cell.font = fontKckh
                else:
                    cell.font = fontKckh2


        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 8.5
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 8.5

        ws.row_dimensions[1].height = 20.25
        ws.row_dimensions[2].height = 20.25
        for rowX in range(3, dataLen + 3):
            ws.row_dimensions[rowX].height = 42.75
